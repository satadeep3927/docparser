import openpyxl


def parse_xlsx(file_path: str) -> str:
    """Convert an XLSX file to Markdown.

    Each worksheet becomes:
        # Sheet Name
        | col | col | ...
        | --- | --- | ...
        | val | val | ...
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Strip fully-empty trailing rows
        while rows and all(cell is None for cell in rows[-1]):
            rows.pop()

        if not rows:
            parts.append(f"# {sheet_name}\n\n*(empty sheet)*")
            continue

        md_table = _rows_to_md_table(rows)
        parts.append(f"# {sheet_name}\n\n{md_table}")

    return "\n\n---\n\n".join(parts)


def _rows_to_md_table(rows) -> str:
    col_count = max(len(r) for r in rows)

    def fmt(cell) -> str:
        if cell is None:
            return ""
        return str(cell).replace("\n", " ").strip()

    normalized = [
        [fmt(cell) for cell in row] + [""] * (col_count - len(row))
        for row in rows
    ]

    header = normalized[0]
    body = normalized[1:]

    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * col_count) + " |")
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)
